import streamlit as st
import requests
import pandas as pd
import os
import io
import base64
from datetime import datetime
import pytz

# =============================================
# 1. ตั้งค่าหน้าตาเว็บ
# =============================================
st.set_page_config(page_title="TSP ERP Issue Tracker", layout="wide")

# =============================================
# 2. โหลด Config (ใช้ secrets เดิม)
# =============================================
try:
    line_bot_token  = st.secrets["LINE_BOT_TOKEN"]
    line_group_id   = st.secrets.get("LINE_GROUP_ID", "")
    tenant_id       = st.secrets["AZURE_TENANT_ID"]
    client_id       = st.secrets["AZURE_CLIENT_ID"]
    client_secret   = st.secrets["AZURE_CLIENT_SECRET"]
    onedrive_folder = st.secrets.get("ONEDRIVE_FOLDER", "Meeting Tracker")
    user_email      = st.secrets.get("ONEDRIVE_USER_EMAIL", "atichat@tspmetal.com")
except Exception:
    line_bot_token  = os.environ.get("LINE_BOT_TOKEN", "")
    line_group_id   = os.environ.get("LINE_GROUP_ID", "")
    tenant_id       = os.environ.get("AZURE_TENANT_ID", "")
    client_id       = os.environ.get("AZURE_CLIENT_ID", "")
    client_secret   = os.environ.get("AZURE_CLIENT_SECRET", "")
    onedrive_folder = os.environ.get("ONEDRIVE_FOLDER", "Meeting Tracker")
    user_email      = os.environ.get("ONEDRIVE_USER_EMAIL", "atichat@tspmetal.com")

TH_TZ = pytz.timezone("Asia/Bangkok")
now   = datetime.now(TH_TZ)

ONEDRIVE_FILE = "ERP_Issue_Log.xlsx"

USERS   = ["ดร.วิทยา", "พี่วรัญ", "อ๊อด"]
BU_LIST = ["ALL", "AM", "OEM", "PM"]
EFFECT  = ["มาก", "ปานกลาง", "น้อย"]
STATUS  = ["Pending", "Considering", "Monitoring", "Closed"]

COLUMNS = [
    "สถานะ", "ลำดับ", "รหัส Issue", "BU", "วันที่รับแจ้ง",
    "ผู้แจ้ง", "หัวข้อปัญหา", "คำอธิบาย", "ระดับผลกระทบ",
    "แนวทางแก้ไข", "วันที่คาดเสร็จ", "ผู้รับผิดชอบ",
    "ช่องทางติดต่อ", "ความคิดเห็น", "ประเภทแถว",
]
# แถวความเห็น/ดำเนินการ จะมี ประเภทแถว = "ความเห็น"
# แถวหลัก จะมี ประเภทแถว = "Issue"

COLOR = {
    "green":      "#1a5c3a",
    "green_lite": "#e8f5ee",
    "green_mid":  "#2d7a56",
    "border":     "#b2d8c8",
    "gold":       "#b8860b",
    "red":        "#c0392b",
    "orange":     "#e67e22",
}

STATUS_COLOR = {
    "Pending":      "#e74c3c",
    "Considering":  "#e67e22",
    "Monitoring":   "#2980b9",
    "Closed":       "#27ae60",
}

# =============================================
# 3. Header
# =============================================
THAI_DAYS   = ["จันทร์","อังคาร","พุธ","พฤหัสบดี","ศุกร์","เสาร์","อาทิตย์"]
THAI_MONTHS = ["","มกราคม","กุมภาพันธ์","มีนาคม","เมษายน","พฤษภาคม","มิถุนายน",
               "กรกฎาคม","สิงหาคม","กันยายน","ตุลาคม","พฤศจิกายน","ธันวาคม"]
day_name   = THAI_DAYS[now.weekday()]
day_num    = now.day
month_name = THAI_MONTHS[now.month]
year_thai  = now.year + 543
time_str   = now.strftime("%H:%M")

logo_b64  = ""
logo_path = os.path.join(os.path.dirname(__file__), "logo.png")
if os.path.exists(logo_path):
    with open(logo_path, "rb") as f:
        logo_b64 = base64.b64encode(f.read()).decode()
logo_html = (f"<img src='data:image/png;base64,{logo_b64}' "
             f"style='width:56px;height:56px;object-fit:contain;border-radius:8px;'/>"
             if logo_b64 else
             f"<div style='width:56px;height:56px;background:{COLOR['green']};border-radius:8px;'></div>")

st.markdown(f"""
<div style="padding:8px 0 16px;">
  <div style="background:{COLOR['green_lite']};border:0.5px solid {COLOR['border']};
              border-radius:12px;padding:14px 24px 0;">
    <div style="display:grid;grid-template-columns:64px 1fr auto;align-items:center;gap:16px;">
      <div>{logo_html}</div>
      <div style="text-align:center;">
        <div style="font-size:26px;font-weight:500;color:{COLOR['green']};line-height:1.2;">TSP Metal Works</div>
        <div style="font-size:16px;color:{COLOR['green_mid']};margin-top:4px;font-weight:500;">
          ระบบแจ้งปัญหา ERP
        </div>
      </div>
      <div style="text-align:right;">
        <div style="font-size:13px;font-weight:500;color:{COLOR['green']};">
          วัน{day_name}ที่ {day_num} {month_name} {year_thai}
        </div>
        <div style="font-size:12px;color:{COLOR['green_mid']};">เวลา {time_str} น.</div>
      </div>
    </div>
    <div style="margin-top:12px;height:3px;
                background:linear-gradient(90deg,{COLOR['green']} 0%,#2d9e6b 50%,{COLOR['green']} 100%);
                border-radius:2px 2px 0 0;"></div>
  </div>
</div>
""", unsafe_allow_html=True)

# =============================================
# 4. OneDrive helpers
# =============================================
def get_token() -> str:
    url  = f"https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token"
    data = {"grant_type": "client_credentials", "client_id": client_id,
            "client_secret": client_secret, "scope": "https://graph.microsoft.com/.default"}
    return requests.post(url, data=data, timeout=15).json()["access_token"]

def upload_onedrive(file_bytes: bytes, filename: str) -> str:
    try:
        token = get_token()
        # อัปโหลดไฟล์
        upload_url = (f"https://graph.microsoft.com/v1.0/users/{user_email}"
                      f"/drive/root:/{onedrive_folder}/{filename}:/content")
        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        }
        r = requests.put(upload_url, headers=headers, data=file_bytes, timeout=30)
        r.raise_for_status()
        item_id = r.json().get("id", "")

        # สร้าง sharing link แบบ "Anyone with the link can view"
        if item_id:
            share_url = (f"https://graph.microsoft.com/v1.0/users/{user_email}"
                         f"/drive/items/{item_id}/createLink")
            share_body = {"type": "view", "scope": "anonymous"}
            sr = requests.post(share_url,
                               headers={**headers, "Content-Type": "application/json"},
                               json=share_body, timeout=15)
            if sr.status_code in (200, 201):
                return sr.json().get("link", {}).get("webUrl", "")

        return r.json().get("webUrl", "")
    except Exception as e:
        st.error(f"❌ Upload OneDrive ไม่สำเร็จ: {e}")
        return ""

def download_onedrive(filename: str) -> bytes:
    try:
        token = get_token()
        url   = (f"https://graph.microsoft.com/v1.0/users/{user_email}"
                 f"/drive/root:/{onedrive_folder}/{filename}:/content")
        r = requests.get(url, headers={"Authorization": f"Bearer {token}"}, timeout=30)
        return r.content if r.status_code == 200 else b""
    except Exception:
        return b""

# =============================================
# 5. LINE helper
# =============================================
def send_line(msg: str):
    if not line_bot_token or not line_group_id:
        return
    try:
        requests.post(
            "https://api.line.me/v2/bot/message/push",
            headers={"Content-Type": "application/json",
                     "Authorization": f"Bearer {line_bot_token}"},
            json={"to": line_group_id, "messages": [{"type": "text", "text": msg}]},
            timeout=10,
        )
    except Exception:
        pass

# =============================================
# 6. โหลดข้อมูลจาก OneDrive
# =============================================
@st.cache_data(ttl=60)
def load_issues() -> pd.DataFrame:
    data = download_onedrive(ONEDRIVE_FILE)
    if data:
        try:
            df = pd.read_excel(io.BytesIO(data))
            return df
        except Exception:
            pass
    return pd.DataFrame(columns=COLUMNS)

def next_issue_id(df: pd.DataFrame) -> str:
    if df.empty or "รหัส Issue" not in df.columns:
        return "ERP-001"
    nums = df["รหัส Issue"].str.extract(r"(\d+)").dropna()[0].astype(int)
    return f"ERP-{(nums.max() + 1):03d}" if len(nums) else "ERP-001"

def next_seq(df: pd.DataFrame) -> int:
    if df.empty or "ลำดับ" not in df.columns:
        return 1
    try:
        return int(df["ลำดับ"].max()) + 1
    except Exception:
        return len(df) + 1

def to_excel_bytes(df: pd.DataFrame) -> bytes:
    from openpyxl.styles import PatternFill, Font, Alignment
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        df.to_excel(w, index=False, sheet_name="ERP Issues")
        ws = w.sheets["ERP Issues"]
        # ปรับความกว้างคอลัมน์
        for col in ws.columns:
            mx = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = max(mx + 4, 12)
        # หา index คอลัมน์ ประเภทแถว
        headers = [c.value for c in ws[1]]
        type_col = headers.index("ประเภทแถว") + 1 if "ประเภทแถว" in headers else None
        # style แถวความเห็น — พื้นหลังเหลืองอ่อน italic
        fill_cmt  = PatternFill("solid", fgColor="FFFDE7")
        fill_issue = PatternFill("solid", fgColor="E8F5EE")
        for row in ws.iter_rows(min_row=2):
            row_type = str(row[type_col-1].value) if type_col else "Issue"
            if row_type == "ความเห็น":
                for cell in row:
                    cell.fill = fill_cmt
                    cell.font = Font(italic=True, size=10)
            else:
                for cell in row:
                    cell.fill = fill_issue
    return buf.getvalue()

# =============================================
# 7. แสดงตาราง issue
# =============================================
df_all = load_issues()

# สรุปตัวเลขตามสถานะ
st.markdown(f"""
<div style="display:flex;gap:12px;flex-wrap:wrap;margin-bottom:16px;">
""" + "".join(
    f"""<div style="background:white;border:1.5px solid {STATUS_COLOR.get(s,'#999')};
                   border-radius:8px;padding:8px 20px;text-align:center;min-width:110px;">
         <div style="font-size:22px;font-weight:700;color:{STATUS_COLOR.get(s,'#999')};">
           {len(df_all[df_all['สถานะ'] == s]) if not df_all.empty and 'สถานะ' in df_all.columns else 0}
         </div>
         <div style="font-size:12px;color:#555;">{s}</div>
       </div>"""
    for s in STATUS
) + "</div>", unsafe_allow_html=True)

# ตารางรายการทั้งหมด
if not df_all.empty:
    with st.expander("📋 ดูรายการทั้งหมด", expanded=False):
        SHOW_COLS = ["รหัส Issue", "BU", "วันที่รับแจ้ง", "ผู้แจ้ง",
                     "หัวข้อปัญหา", "ระดับผลกระทบ", "สถานะ",
                     "ผู้รับผิดชอบ", "วันที่คาดเสร็จ"]
        show = [c for c in SHOW_COLS if c in df_all.columns]
        rows_html = ""
        for i, (_, row) in enumerate(df_all[show].iterrows()):
            bg = COLOR["green_lite"] if i % 2 == 0 else "white"
            cells = ""
            for c in show:
                v = str(row.get(c, ""))
                v = "—" if v in ("nan", "None", "NaT", "") else v
                if c == "สถานะ":
                    sc = STATUS_COLOR.get(v, "#999")
                    cells += (f"<td style='padding:5px 8px;font-size:12px;'>"
                              f"<span style='background:{sc};color:white;padding:2px 8px;"
                              f"border-radius:10px;font-size:11px;'>{v}</span></td>")
                else:
                    cells += f"<td style='padding:5px 8px;font-size:12px;'>{v}</td>"
            rows_html += f"<tr style='background:{bg}'>{cells}</tr>"

        hdr = "".join(
            f"<th style='padding:6px 8px;background:{COLOR['green']};color:white;"
            f"font-size:12px;text-align:left;white-space:nowrap;'>{c}</th>"
            for c in show
        )
        st.markdown(
            f"<div style='overflow-x:auto'><table style='border-collapse:collapse;width:100%;'>"
            f"<thead><tr>{hdr}</tr></thead><tbody>{rows_html}</tbody></table></div>",
            unsafe_allow_html=True
        )

# =============================================
# 8. ฟอร์มแจ้งปัญหาใหม่
# =============================================
st.divider()
st.subheader("➕ แจ้งปัญหา ERP ใหม่")

with st.form("issue_form", clear_on_submit=True):
    c1, c2, c3 = st.columns([2, 2, 2])
    with c1:
        informer = st.selectbox("ผู้แจ้ง *", USERS)
    with c2:
        bu = st.selectbox("BU *", BU_LIST)
    with c3:
        effect = st.selectbox("ระดับผลกระทบ *", EFFECT)

    topic = st.text_input("หัวข้อปัญหา *", placeholder="สั้น กระชับ เช่น ราคาสินค้าไม่อัปเดต")
    desc  = st.text_area("คำอธิบายรายละเอียด", height=100,
                         placeholder="อธิบายปัญหา ขั้นตอนที่ทำ และผลที่เกิดขึ้น...")

    c4, c5 = st.columns([3, 1])
    with c4:
        solution = st.text_area("แนวทางแก้ไข (ถ้ามี)", height=80, placeholder="เช่น รอ patch จาก vendor")
    with c5:
        due_date   = st.date_input("วันที่คาดเสร็จ", value=None)
        status_new = st.selectbox("สถานะ", STATUS)

    c6, c7 = st.columns(2)
    with c6:
        incharge = st.text_input("ผู้รับผิดชอบ", placeholder="เช่น ทีม IT, อ๊อด")
    with c7:
        contact  = st.text_input("ช่องทางติดต่อ", placeholder="เช่น email, LINE")

    comment = st.text_area("ความคิดเห็นเพิ่มเติม", height=70)

    submitted = st.form_submit_button("💾 บันทึกและแจ้ง LINE", type="primary",
                                       use_container_width=True)

if submitted:
    if not topic.strip():
        st.warning("⚠️ กรุณากรอกหัวข้อปัญหาครับ")
    else:
        with st.spinner("💾 กำลังบันทึก..."):
            st.cache_data.clear()
            df = load_issues()
            issue_id = next_issue_id(df)
            seq      = next_seq(df)
            due_str  = due_date.strftime("%d/%m/%Y") if due_date else "ไม่มีกำหนดเสร็จที่แน่นอน"

            new_row = {
                "สถานะ":         status_new,
                "ลำดับ":          seq,
                "รหัส Issue":     issue_id,
                "BU":             bu,
                "วันที่รับแจ้ง":  now.strftime("%d/%m/%Y %H:%M"),
                "ผู้แจ้ง":        informer,
                "หัวข้อปัญหา":    topic.strip(),
                "คำอธิบาย":       desc.strip(),
                "ระดับผลกระทบ":   effect,
                "แนวทางแก้ไข":    solution.strip(),
                "วันที่คาดเสร็จ": due_str,
                "ผู้รับผิดชอบ":   incharge.strip(),
                "ช่องทางติดต่อ":  contact.strip(),
                "ความคิดเห็น":    comment.strip(),
                "ประเภทแถว":      "Issue",
            }

            df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)

            # อัปโหลด OneDrive
            link = upload_onedrive(to_excel_bytes(df), ONEDRIVE_FILE)
            if link:
                st.success(f"✅ บันทึกสำเร็จ! [{issue_id}] {topic.strip()}")
                st.markdown(f"📄 [เปิดไฟล์ ERP Issue Log]({link})")
            else:
                st.warning("⚠️ บันทึก OneDrive ไม่สำเร็จ แต่ข้อมูลส่ง LINE แล้ว")

            # แจ้ง LINE
            effect_icon = {"มาก": "🔴", "ปานกลาง": "🟡", "น้อย": "🟢"}.get(effect, "⚪")
            line_msg = (
                f"🔔 แจ้งปัญหา ERP ใหม่ — TSP Metal Works\n"
                f"วันที่ {now.strftime('%d/%m/%Y')} เวลา {now.strftime('%H:%M')} น.\n"
                f"{'─'*30}\n"
                f"🆔 {issue_id}  |  BU: {bu}\n"
                f"👤 ผู้แจ้ง: {informer}\n"
                f"📌 {topic.strip()}\n"
            )
            if desc.strip():
                line_msg += f"📝 {desc.strip()[:120]}{'...' if len(desc.strip()) > 120 else ''}\n"
            line_msg += (
                f"{effect_icon} ผลกระทบ: {effect}\n"
                f"📊 สถานะ: {status_new}\n"
                f"🎯 กำหนดเสร็จ: {due_str}\n"
            )
            if incharge.strip():
                line_msg += f"👷 ผู้รับผิดชอบ: {incharge.strip()}\n"
            if link:
                line_msg += f"📄 ดูรายละเอียด: {link}"

            send_line(line_msg)
            st.success("📲 ส่ง LINE กลุ่มเรียบร้อยแล้วครับ!")
            st.cache_data.clear()

# =============================================
# 9. อัปเดตสถานะ issue ที่มีอยู่
# =============================================
st.divider()
st.subheader("✏️ อัปเดตสถานะ / ความคิดเห็น")

df_active = load_issues()
open_issues = df_active[
    df_active.get("สถานะ", pd.Series(dtype=str)).isin(["Pending", "Considering", "Monitoring"])
] if not df_active.empty and "สถานะ" in df_active.columns else pd.DataFrame()

if open_issues.empty:
    st.info("ไม่มี issue ที่ยังเปิดอยู่ครับ 🎉")
else:
    for idx, (_, row) in enumerate(open_issues.iterrows()):
        iid     = str(row.get("รหัส Issue", ""))
        itopic  = str(row.get("หัวข้อปัญหา", ""))
        istatus = str(row.get("สถานะ", "Pending"))
        ibu     = str(row.get("BU", ""))
        idue    = str(row.get("วันที่คาดเสร็จ", "—"))
        sc      = STATUS_COLOR.get(istatus, "#999")

        label = (f"**{iid}** — {itopic}  "
                 f"| BU: {ibu}  | 🎯 {idue}")

        with st.expander(label, expanded=False):
            st.markdown(
                f"<span style='background:{sc};color:white;padding:2px 10px;"
                f"border-radius:10px;font-size:12px;'>{istatus}</span>",
                unsafe_allow_html=True
            )
            st.markdown("")

            cu1, cu2, cu3 = st.columns([2, 3, 1])
            with cu1:
                commenter = st.selectbox(
                    "ผู้ดำเนินการ / ผู้ให้ความเห็น:",
                    USERS, key=f"who_{idx}_{iid}"
                )
            with cu2:
                new_comment = st.text_area(
                    "ความคิดเห็น / การดำเนินการ:",
                    key=f"cmt_{idx}_{iid}", height=80,
                    placeholder="เช่น ติดต่อ vendor แล้ว รอ patch\nทดสอบระบบแล้ว ปกติ"
                )
            with cu3:
                new_status = st.selectbox(
                    "สถานะ:", STATUS,
                    index=STATUS.index(istatus) if istatus in STATUS else 0,
                    key=f"st_{idx}_{iid}"
                )

            # บันทึกลง session_state ทันทีที่ widget เปลี่ยน
            st.session_state[f"saved_who_{iid}"]    = commenter
            st.session_state[f"saved_cmt_{iid}"]    = new_comment
            st.session_state[f"saved_status_{iid}"] = new_status

            if st.button(f"💾 บันทึก {iid}", key=f"save_{idx}_{iid}"):
                _who     = st.session_state.get(f"saved_who_{iid}",    USERS[0])
                _comment = st.session_state.get(f"saved_cmt_{iid}",    "")
                _status  = st.session_state.get(f"saved_status_{iid}", istatus)

                with st.spinner("บันทึก..."):
                    st.cache_data.clear()
                    df_cur = load_issues()
                    mask   = df_cur["รหัส Issue"] == iid

                    if mask.any():
                        # อัปเดตสถานะในแถว Issue หลัก
                        df_cur.loc[mask & (df_cur["ประเภทแถว"] == "Issue"), "สถานะ"] = _status

                        # เพิ่มแถวความเห็นใหม่ต่อท้าย (ถ้ามีข้อความ)
                        if _comment.strip():
                            cmt_row = {
                                "สถานะ":        _status,
                                "ลำดับ":         "",
                                "รหัส Issue":    iid,
                                "BU":            str(row.get("BU", "")),
                                "วันที่รับแจ้ง": now.strftime("%d/%m/%Y %H:%M"),
                                "ผู้แจ้ง":       _who,
                                "หัวข้อปัญหา":   "",
                                "คำอธิบาย":      "",
                                "ระดับผลกระทบ":  "",
                                "แนวทางแก้ไข":   "",
                                "วันที่คาดเสร็จ":"",
                                "ผู้รับผิดชอบ":  "",
                                "ช่องทางติดต่อ": "",
                                "ความคิดเห็น":   _comment.strip(),
                                "ประเภทแถว":     "ความเห็น",
                            }
                            df_cur = pd.concat([df_cur, pd.DataFrame([cmt_row])],
                                               ignore_index=True)

                        link = upload_onedrive(to_excel_bytes(df_cur), ONEDRIVE_FILE)
                        if link:
                            st.success(f"✅ อัปเดต {iid} เรียบร้อยแล้ว")
                        else:
                            st.error("❌ Upload OneDrive ไม่สำเร็จ")

                        if _status != istatus or _comment.strip():
                            line_msg = (
                                f"🔔 อัปเดตปัญหา ERP — TSP Metal Works\n"
                                f"วันที่ {now.strftime('%d/%m/%Y')} เวลา {now.strftime('%H:%M')} น.\n"
                                f"{'─'*30}\n"
                                f"🆔 {iid} — {itopic}\n"
                            )
                            if _status != istatus:
                                line_msg += f"📊 สถานะ: {istatus} → {_status}\n"
                            if _comment.strip():
                                line_msg += f"💬 {_who}: {_comment.strip()[:120]}\n"
                            if link:
                                line_msg += f"📄 ดูรายละเอียด: {link}"
                            send_line(line_msg)
                            st.success("📲 แจ้ง LINE แล้วครับ!")

                        st.session_state.pop(f"saved_who_{iid}",    None)
                        st.session_state.pop(f"saved_cmt_{iid}",    None)
                        st.session_state.pop(f"saved_status_{iid}", None)
                        st.cache_data.clear()
                        st.rerun()

# =============================================
# 10. เครื่องมือหา LINE Group ID
# =============================================
st.divider()
with st.expander("🔍 ค้นหา LINE Group ID (สำหรับตั้งค่าครั้งแรก)", expanded=False):
    st.markdown("""
**วิธีใช้:**
1. เชิญ LINE Bot เข้ากลุ่มที่ต้องการก่อน
2. กดปุ่ม **"รับ Webhook"** ด้านล่าง
3. ไปที่กลุ่ม LINE แล้วพิมพ์ข้อความอะไรก็ได้ 1 ข้อความ ภายใน 60 วินาที
4. กด **"ตรวจสอบ Group ID"** เพื่อดูผล
""")

    col_a, col_b = st.columns(2)

    with col_a:
        if st.button("📡 ดึง Group ID จาก Recent Messages", use_container_width=True):
            if not line_bot_token:
                st.error("ไม่พบ LINE_BOT_TOKEN ใน secrets")
            else:
                try:
                    # ดึง recent messages ผ่าน LINE Messaging API
                    # วิธีที่ทำงานได้จริงคือ GET /v2/bot/insight/message/delivery
                    # แต่วิธีที่ง่ายกว่าคือส่ง test message แล้วดู response
                    headers = {
                        "Authorization": f"Bearer {line_bot_token}",
                        "Content-Type": "application/json"
                    }
                    # ดึงข้อมูล bot profile เพื่อตรวจสอบ token ก่อน
                    r = requests.get("https://api.line.me/v2/bot/info",
                                     headers=headers, timeout=10)
                    if r.status_code == 200:
                        bot_info = r.json()
                        st.success(f"✅ Token ถูกต้อง — Bot: **{bot_info.get('displayName', '')}**")
                        st.info("👆 เชิญ bot นี้เข้ากลุ่ม แล้วใช้วิธีด้านล่างเพื่อหา Group ID")
                    else:
                        st.error(f"Token ไม่ถูกต้อง: {r.status_code}")
                except Exception as e:
                    st.error(f"Error: {e}")

    with col_b:
        group_id_input = st.text_input("ทดสอบส่งข้อความไปกลุ่ม (ใส่ Group ID ที่รู้แล้ว):",
                                        placeholder="Cxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx")
        if st.button("📤 ทดสอบส่ง", use_container_width=True):
            if group_id_input.strip():
                try:
                    headers = {
                        "Authorization": f"Bearer {line_bot_token}",
                        "Content-Type": "application/json"
                    }
                    payload = {
                        "to": group_id_input.strip(),
                        "messages": [{"type": "text",
                                      "text": "✅ ทดสอบการเชื่อมต่อ TSP ERP Issue Tracker"}]
                    }
                    r = requests.post("https://api.line.me/v2/bot/message/push",
                                      headers=headers, json=payload, timeout=10)
                    if r.status_code == 200:
                        st.success("✅ ส่งสำเร็จ! Group ID นี้ถูกต้อง นำไปใส่ใน secrets ได้เลย")
                        st.code(f'LINE_GROUP_ID = "{group_id_input.strip()}"')
                    else:
                        st.error(f"ส่งไม่สำเร็จ: {r.status_code} — {r.text}")
                except Exception as e:
                    st.error(f"Error: {e}")
            else:
                st.warning("กรุณาใส่ Group ID ก่อนครับ")

    st.markdown("---")
    st.markdown("**วิธีหา Group ID แบบง่ายที่สุด:**")
    st.markdown("""
1. ไปที่ [LINE Developers Console](https://developers.line.biz/console/)
2. เข้า channel → แท็บ **Messaging API**
3. เปิด **Use webhooks** → ตั้ง Webhook URL เป็น `https://webhook.site/` (ชั่วคราว)
4. เชิญ bot เข้ากลุ่ม แล้วพิมพ์ข้อความในกลุ่ม
5. ดูใน webhook.site จะเห็น `"groupId": "Cxxxxxxxxx"` ในข้อมูล JSON
""")
